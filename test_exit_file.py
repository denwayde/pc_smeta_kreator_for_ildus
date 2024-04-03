from openpyxl import load_workbook
import asyncio
# arr1 = [
#     [112, 'ул Богдана Хмельницкого', 'д. 52', 1, 'ТП-159/1-ф7'],
#     [113, 'ул Грибоедова', 'д. 4', 'кв. 2', 1, 'ТП-237/1-ф1'],
#     [114, 'ул Железнодорожная', 'д. 12', 1, 'ТП-35/1-ф4'],
#     [115, 'ул Кирова', 'д. 40', 'кв. 2', 3, 'ТП-12'],
#     [116, 'ул Чертыгашева', 'д. 11', 'кв. 3', 1, 'ТП-564/1-ф4'],
#     [117, 'ул Сурикова', 'д. 4', 3, 'ТП-12'],
#     [118, 'ул Кирова', 'д. 55', 1, 'ТП-12/1-ф3'],
#     [119, 'ул Гайдара', 'д. 6', 3, 'ТП-237/1-ф2'],
#     [120, 'ул Сурикова', 'д. 18', 'кв. 2', 1, 'ТП-12/1-ф4'],
#     [121, 'ул Ботаническая', 'д. 29', 'кв. 2', 1, 'ТП-159/1-ф23'],
# ]

# arr1 = [
#     [112, 'ул Богдана Хмельницкого', 'д. 52', 1, 'ТП-159/1-ф7', '194'], [113, 'ул Грибоедова', 'д. 4', 'кв. 2', 1, 'ТП-237/1-ф1', '194'], [114, 'ул Железнодорожная', 'д. 12', 1, 'ТП-35/1-ф4', '194'], [115, 'ул Кирова', 'д. 40', 'кв. 2', 3, 'ТП-12', '194'], [116, 'ул Чертыгашева', 'д. 11', 'кв. 3', 1, 'ТП-564/1-ф4', '194'], [117, 'ул Сурикова', 'д. 4', 3, 'ТП-12', '194'], [118, 'ул Кирова', 'д. 55', 1, 'ТП-12/1-ф3', '194'], [119, 'ул Гайдара', 'д. 6', 3, 'ТП-237/1-ф2', '194'], [120, 'ул Сурикова', 'д. 18', 'кв. 2', 1, 'ТП-12/1-ф4', '194'], [121, 'ул Ботаническая', 'д. 29', 'кв. 2', 1, 'ТП-159/1-ф23', '194']
# ]

def address(new_ad, new_ap):
    s = f"Объект - Установка коммерческого учета электрической энергии для электроснабжения электроустановок жилого дома, расположенного по адресу: г. Абакан, {new_ad} (ВЛ-0,4кВ от {new_ap})"
    return s

async def file_creator(arr1):
    for x in arr1:
        wb = None
        if x[-3] == 1:
            wb = load_workbook('1f.xlsx')
        elif x[-3] == 3:
            wb = load_workbook('3f.xlsx')
        # Выбираем активный лист
        sheet = wb.active
        # Читаем значение из ячейк
        value_B10 = sheet['B10'].value
        value_I18 = sheet['I18'].value
        value_M12 = sheet['M12'].value
        street_name = str(x[1]) + ", " + str(x[2]).replace('/', ' корп. ') + (", "+ x[3]+" " if isinstance(x[3], str) else '')
        index2 = value_M12.find('/')
        new_string2 = x[-1] + value_M12[index2:]
        end_str = str(x[-2]).replace('/', '')
        # Изменяем значения в выбранных ячейках
        sheet['B10'] = address(street_name, x[-2])
        sheet['I18'] = x[0]
        sheet['M12'] = new_string2
        # Сохраняем изменения в новый файл
        wb.save('выходные файлы/'+str(x[0])+" "+street_name+" "+str(x[-3])+" "+end_str+'.xlsx')
        await asyncio.sleep(0.01)
        

# asyncio.run(file_creator(arr1))